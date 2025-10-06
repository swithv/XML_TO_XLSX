"""
Módulo responsável pela exportação de dados para Excel
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import config
from utils.logger import logger

class ExcelExporter:
    """Exporta DataFrames para arquivos Excel formatados"""
    
    def __init__(self):
        self.config = config.EXCEL_CONFIG
    
    def export_to_excel(self, df: pd.DataFrame, filename: str = None) -> BytesIO:
        """
        Exporta DataFrame para Excel formatado
        
        Args:
            df: DataFrame a exportar
            filename: Nome do arquivo (opcional)
            
        Returns:
            BytesIO com o conteúdo do Excel
        """
        if df.empty:
            logger.warning("DataFrame vazio, nada para exportar")
            return None
        
        # Cria buffer em memória
        output = BytesIO()
        
        try:
            # Exporta para Excel
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Dados')
                
                # Acessa o workbook para aplicar formatações
                workbook = writer.book
                worksheet = writer.sheets['Dados']
                
                # Aplica formatações
                self._format_header(worksheet)
                self._format_columns(worksheet, df)
                self._adjust_column_widths(worksheet, df)
                self._apply_borders(worksheet, len(df))
            
            output.seek(0)
            logger.info(f"Excel exportado com sucesso: {len(df)} linhas")
            return output
        
        except Exception as e:
            logger.error(f"Erro ao exportar Excel: {e}")
            return None
    
    def _format_header(self, worksheet):
        """
        Formata o cabeçalho da planilha
        
        Args:
            worksheet: Worksheet do openpyxl
        """
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Aplica formatação na primeira linha (cabeçalho)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _format_columns(self, worksheet, df: pd.DataFrame):
        """
        Formata colunas baseado em seu tipo de dado
        
        Args:
            worksheet: Worksheet do openpyxl
            df: DataFrame original
        """
        for idx, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(idx)
            
            # Formata colunas monetárias
            if self._is_monetary_column(column):
                for row in range(2, len(df) + 2):
                    cell = worksheet[f'{column_letter}{row}']
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            # Formata colunas de data
            elif 'data' in column.lower() or pd.api.types.is_datetime64_any_dtype(df[column]):
                for row in range(2, len(df) + 2):
                    cell = worksheet[f'{column_letter}{row}']
                    cell.number_format = 'DD/MM/YYYY HH:MM'
                    cell.alignment = Alignment(horizontal='center')
            
            # Centraliza colunas de documento (CPF/CNPJ)
            elif 'cnpj' in column.lower() or 'cpf' in column.lower():
                for row in range(2, len(df) + 2):
                    cell = worksheet[f'{column_letter}{row}']
                    cell.alignment = Alignment(horizontal='center')
    
    def _adjust_column_widths(self, worksheet, df: pd.DataFrame):
        """
        Ajusta largura das colunas automaticamente
        
        Args:
            worksheet: Worksheet do openpyxl
            df: DataFrame original
        """
        for idx, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(idx)
            
            # Calcula largura baseada no conteúdo
            max_length = len(str(column))
            
            # Verifica os valores da coluna (primeiras 100 linhas)
            for value in df[column].head(100):
                try:
                    if len(str(value)) > max_length:
                        max_length = len(str(value))
                except:
                    pass
            
            # Define largura (mínimo 12, máximo 50)
            adjusted_width = min(max(max_length + 2, 12), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_borders(self, worksheet, num_rows: int):
        """
        Aplica bordas nas células
        
        Args:
            worksheet: Worksheet do openpyxl
            num_rows: Número de linhas de dados
        """
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplica bordas em todas as células com dados
        for row in worksheet.iter_rows(min_row=1, max_row=num_rows + 1):
            for cell in row:
                cell.border = thin_border
    
    def _is_monetary_column(self, column_name: str) -> bool:
        """Verifica se coluna contém valores monetários"""
        monetary_keywords = ['valor', 'total', 'preco', 'preço', 'custo', 'desconto']
        return any(keyword in column_name.lower() for keyword in monetary_keywords)
    
    def export_with_summary(self, df: pd.DataFrame) -> BytesIO:
        """
        Exporta Excel com uma aba de resumo adicional
        
        Args:
            df: DataFrame a exportar
            
        Returns:
            BytesIO com o Excel
        """
        if df.empty:
            return None
        
        output = BytesIO()
        
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Aba de dados
                df.to_excel(writer, index=False, sheet_name='Dados')
                
                # Aba de resumo
                summary_data = self._create_summary(df)
                summary_df = pd.DataFrame(list(summary_data.items()), 
                                         columns=['Métrica', 'Valor'])
                summary_df.to_excel(writer, index=False, sheet_name='Resumo')
                
                # Formata ambas as abas
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    self._format_header(worksheet)
                    self._adjust_column_widths(worksheet, 
                                              df if sheet_name == 'Dados' else summary_df)
            
            output.seek(0)
            return output
        
        except Exception as e:
            logger.error(f"Erro ao exportar Excel com resumo: {e}")
            return None
    
    def _create_summary(self, df: pd.DataFrame) -> dict:
        """
        Cria resumo estatístico do DataFrame
        
        Args:
            df: DataFrame
            
        Returns:
            Dicionário com métricas
        """
        summary = {
            'Total de Registros': len(df),
            'Data de Geração': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        }
        
        # Adiciona soma de colunas monetárias
        for column in df.columns:
            if self._is_monetary_column(column) and pd.api.types.is_numeric_dtype(df[column]):
                summary[f'Total {column}'] = df[column].sum()
        
        return summary